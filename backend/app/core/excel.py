from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from io import BytesIO
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, ZipFile, is_zipfile

from defusedxml.common import DefusedXmlException  # type: ignore[import-untyped]
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from pydantic import BaseModel, ValidationError

from app.core.exceptions import AppError

MAX_XLSX_BYTES = 10 * 1024 * 1024
MAX_XLSX_DATA_ROWS = 10_000


@dataclass(frozen=True)
class ExcelIssue:
    worksheet: str | None
    row: int | None
    column: str | None
    field: str | None
    message: str

    def as_dict(self) -> dict[str, str | int | None]:
        return {
            "worksheet": self.worksheet,
            "row": self.row,
            "column": self.column,
            "field": self.field,
            "message": self.message,
        }


@dataclass(frozen=True)
class ExcelRow:
    worksheet: str
    row: int
    value: BaseModel


@dataclass(frozen=True)
class ExcelReadResult:
    rows: list[ExcelRow]
    issues: list[ExcelIssue]


class ExcelValidationError(AppError):
    status_code = 422

    def __init__(self, issues: Iterable[ExcelIssue]) -> None:
        self.issues = list(issues)
        super().__init__(
            {
                "message": "Excel validation failed",
                "issues": [issue.as_dict() for issue in self.issues],
            }
        )


def _headers(model_type: type[BaseModel]) -> list[tuple[str, str]]:
    return [
        (field_name, field.alias or field_name)
        for field_name, field in model_type.model_fields.items()
    ]


def _header_value(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _is_blank(row: tuple[object, ...]) -> bool:
    return not any(value is not None and str(value).strip() for value in row)


def _workbook_issue(message: str, *, worksheet: str | None = None) -> ExcelReadResult:
    return ExcelReadResult(
        rows=[],
        issues=[
            ExcelIssue(
                worksheet=worksheet,
                row=None,
                column=None,
                field=None,
                message=message,
            )
        ],
    )


def load_xlsx_workbook(content: bytes, *, max_bytes: int = MAX_XLSX_BYTES) -> Workbook:
    if len(content) > max_bytes:
        raise ExcelValidationError(
            [
                ExcelIssue(
                    worksheet=None,
                    row=None,
                    column=None,
                    field=None,
                    message=f"Workbook exceeds the {max_bytes} byte limit",
                )
            ]
        )
    if not is_zipfile(BytesIO(content)):
        raise ExcelValidationError(
            [ExcelIssue(None, None, None, None, "File is not a valid XLSX workbook")]
        )
    try:
        with ZipFile(BytesIO(content)) as archive:
            if any(
                name.casefold().endswith("vbaproject.bin")
                for name in archive.namelist()
            ):
                raise ExcelValidationError(
                    [
                        ExcelIssue(
                            None,
                            None,
                            None,
                            None,
                            "Macro-enabled workbooks are not supported",
                        )
                    ]
                )
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except ExcelValidationError:
        raise
    except (
        BadZipFile,
        DefusedXmlException,
        KeyError,
        OSError,
        ParseError,
        ValueError,
    ) as error:
        raise ExcelValidationError(
            [ExcelIssue(None, None, None, None, "File is not a valid XLSX workbook")]
        ) from error


def read_xlsx_rows[ModelT: BaseModel](
    content: bytes,
    *,
    model_type: type[ModelT],
    worksheet_name: str,
    max_bytes: int = MAX_XLSX_BYTES,
    max_data_rows: int = MAX_XLSX_DATA_ROWS,
) -> ExcelReadResult:
    try:
        workbook = load_xlsx_workbook(content, max_bytes=max_bytes)
    except ExcelValidationError as error:
        return ExcelReadResult(rows=[], issues=error.issues)
    try:
        if worksheet_name not in workbook.sheetnames:
            return _workbook_issue(f"Worksheet '{worksheet_name}' was not found")
        worksheet = workbook[worksheet_name]
        row_iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(row_iterator)
        except StopIteration:
            return _workbook_issue("Worksheet is empty", worksheet=worksheet_name)

        headers = _headers(model_type)
        aliases = {alias: field_name for field_name, alias in headers}
        columns: dict[str, int] = {}
        issues: list[ExcelIssue] = []
        for column_index, value in enumerate(header_row):
            header = _header_value(value)
            if header not in aliases:
                continue
            if header in columns:
                issues.append(
                    ExcelIssue(
                        worksheet=worksheet_name,
                        row=1,
                        column=header,
                        field=aliases[header],
                        message="Duplicate declared header",
                    )
                )
            else:
                columns[header] = column_index
        for field_name, alias in headers:
            if alias not in columns:
                issues.append(
                    ExcelIssue(
                        worksheet=worksheet_name,
                        row=1,
                        column=alias,
                        field=field_name,
                        message="Missing declared header",
                    )
                )
        if issues:
            return ExcelReadResult(rows=[], issues=issues)

        rows: list[ExcelRow] = []
        data_rows = 0
        for row_number, values in enumerate(row_iterator, start=2):
            if _is_blank(values):
                continue
            data_rows += 1
            if data_rows > max_data_rows:
                return ExcelReadResult(
                    rows=[],
                    issues=[
                        ExcelIssue(
                            worksheet=worksheet_name,
                            row=row_number,
                            column=None,
                            field=None,
                            message=(
                                f"Workbook exceeds the {max_data_rows} data row limit"
                            ),
                        )
                    ],
                )
            payload = {
                alias: values[column_index] if column_index < len(values) else None
                for alias, column_index in columns.items()
            }
            try:
                model = model_type.model_validate(payload, by_alias=True, by_name=False)
            except ValidationError as error:
                for item in error.errors():
                    location = item["loc"][0] if item["loc"] else None
                    error_field: str | None = aliases.get(
                        str(location), str(location) if location else None
                    )
                    error_alias: str | None = next(
                        (
                            header
                            for header, name in aliases.items()
                            if name == error_field
                        ),
                        None,
                    )
                    issues.append(
                        ExcelIssue(
                            worksheet=worksheet_name,
                            row=row_number,
                            column=error_alias,
                            field=error_field,
                            message=str(item["msg"]),
                        )
                    )
                continue
            rows.append(ExcelRow(worksheet=worksheet_name, row=row_number, value=model))
        return ExcelReadResult(rows=rows, issues=issues)
    finally:
        workbook.close()


def create_xlsx(
    rows: Iterable[BaseModel], *, model_type: type[BaseModel], worksheet_name: str
) -> BytesIO:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(worksheet_name)
    field_names = list(model_type.model_fields)
    worksheet.append([alias for _, alias in _headers(model_type)])
    for row in rows:
        values = row.model_dump(by_alias=True)
        worksheet.append(
            [
                values.get(model_type.model_fields[name].alias or name)
                for name in field_names
            ]
        )
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
