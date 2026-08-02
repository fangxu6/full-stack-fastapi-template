import hashlib
import json
import re
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import cast

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.utils.datetime import from_excel  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]
from pydantic import ValidationError
from sqlmodel import Session, select

from app.core.audit import (
    AUDIT_ACTOR_SESSION_KEY,
    bind_audit_actor,
    clear_audit_actor,
)
from app.core.excel import (
    MAX_XLSX_DATA_ROWS,
    ExcelIssue,
    ExcelRow,
    ExcelValidationError,
    load_xlsx_workbook,
    read_xlsx_rows,
)
from app.core.exceptions import BadRequestError, ConflictError
from app.models import (
    InventoryDocument,
    InventoryDocumentLine,
    InventoryImportBatch,
    InventoryLedgerEntry,
    LegacyImportRow,
    ProcessingUnit,
    ReceivingUnit,
    User,
)
from app.models.inventory import (
    InventoryDocumentType,
    InventoryLedgerKind,
    InventoryMovementType,
    LegacyWorkbookKind,
)
from app.modules.inventory import service
from app.schemas.inventory import (
    InventoryDocumentCreate,
    InventoryDocumentExcelRow,
    InventoryExcelImportPublic,
    InventoryLineCreate,
    LegacyInventoryExcelImportPublic,
)

IMPORTER_VERSION = "inventory-xlsx-v2"
DOCUMENT_WORKSHEET_NAME = "单据导入"
MISSING_ITEM_CODE = "未填写品号"
MISSING_WOOL_CONTENT = "未填写含毛量"
MISSING_DYE_LOT = "未分缸"


def _document_issue(row: ExcelRow, *, field_name: str, message: str) -> ExcelIssue:
    field = InventoryDocumentExcelRow.model_fields[field_name]
    return ExcelIssue(
        worksheet=row.worksheet,
        row=row.row,
        column=field.alias or field_name,
        field=field_name,
        message=message,
    )


def import_document_workbook(
    *, session: Session, content: bytes
) -> InventoryExcelImportPublic:
    result = read_xlsx_rows(
        content,
        model_type=InventoryDocumentExcelRow,
        worksheet_name=DOCUMENT_WORKSHEET_NAME,
    )
    if result.issues:
        raise ExcelValidationError(result.issues)

    groups: dict[str, list[ExcelRow]] = {}
    for row in result.rows:
        value = cast(InventoryDocumentExcelRow, row.value)
        groups.setdefault(value.document_number, []).append(row)

    issues: list[ExcelIssue] = []
    document_fields = (
        "document_type",
        "business_date",
        "processing_unit_name",
        "receiving_unit_name",
        "remarks",
    )
    for group in groups.values():
        first = cast(InventoryDocumentExcelRow, group[0].value)
        for row in group[1:]:
            value = cast(InventoryDocumentExcelRow, row.value)
            for field_name in document_fields:
                if getattr(value, field_name) != getattr(first, field_name):
                    issues.append(
                        _document_issue(
                            row,
                            field_name=field_name,
                            message="Document fields must match within a document number",
                        )
                    )
    if issues:
        raise ExcelValidationError(issues)

    created_numbers: list[str] = []
    for number, group in groups.items():
        first = cast(InventoryDocumentExcelRow, group[0].value)
        try:
            processing_unit_id = service.resolve_active_unit_name(
                session=session,
                model=ProcessingUnit,
                name=first.processing_unit_name,
            )
        except BadRequestError as error:
            issues.append(
                _document_issue(
                    group[0], field_name="processing_unit_name", message=str(error)
                )
            )
            continue
        receiving_unit_id = None
        if first.receiving_unit_name:
            try:
                receiving_unit_id = service.resolve_active_unit_name(
                    session=session,
                    model=ReceivingUnit,
                    name=first.receiving_unit_name,
                )
            except BadRequestError as error:
                issues.append(
                    _document_issue(
                        group[0],
                        field_name="receiving_unit_name",
                        message=str(error),
                    )
                )
                continue
        try:
            document_in = InventoryDocumentCreate(
                document_type=first.document_type,
                business_date=first.business_date,
                processing_unit_id=processing_unit_id,
                receiving_unit_id=receiving_unit_id,
                document_number=number,
                remarks=first.remarks,
                lines=[
                    InventoryLineCreate(
                        item_name=value.item_name,
                        item_code=value.item_code,
                        wool_content=value.wool_content,
                        color_code=value.color_code,
                        dye_lot_no=value.dye_lot_no,
                        quantity_rolls=value.quantity_rolls,
                        quantity_meters=value.quantity_meters,
                    )
                    for row in group
                    for value in [cast(InventoryDocumentExcelRow, row.value)]
                ],
            )
        except ValidationError as error:
            issues.append(
                _document_issue(
                    group[0], field_name="document_number", message=str(error)
                )
            )
            continue
        try:
            with session.begin_nested():
                service.create_document(session=session, document_in=document_in)
        except (BadRequestError, ConflictError) as error:
            issues.append(
                _document_issue(
                    group[0], field_name="document_number", message=str(error)
                )
            )
            continue
        created_numbers.append(number)
    if issues:
        raise ExcelValidationError(issues)
    return InventoryExcelImportPublic(
        created_documents=len(created_numbers), document_numbers=created_numbers
    )


def _sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _text(value: object, placeholder: str | None = None) -> str | None:
    if value is None or not str(value).strip():
        return placeholder
    return str(value).strip()


def _number(value: object) -> Decimal:
    try:
        return Decimal(str(value or 0).replace(",", ""))
    except Exception as err:
        raise BadRequestError(f"Invalid inventory quantity: {value}") from err


def _date(value: object, *, epoch: datetime) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value or "").strip()
    year_closing_match = re.fullmatch(r"(\d{4})年结存", text)
    if year_closing_match:
        return date(int(year_closing_match.group(1)), 12, 31)

    try:
        numeric_value = Decimal(text)
    except Exception:
        numeric_value = None
    if numeric_value is not None and 30_000 <= numeric_value <= 100_000:
        converted = from_excel(float(numeric_value), epoch)
        return converted.date() if isinstance(converted, datetime) else converted

    slash_date_match = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", text)
    if slash_date_match:
        return date(
            int(slash_date_match.group(1)),
            int(slash_date_match.group(2)),
            int(slash_date_match.group(3)),
        )

    try:
        return datetime.fromisoformat(text.replace("/", "-")).date()
    except ValueError as err:
        raise BadRequestError(f"Invalid inventory date: {value}") from err


def _header_key(value: object) -> str:
    return "".join(str(value or "").strip().lower().split())


def _row_values(sheet: Worksheet) -> list[tuple[int, dict[str, object]]]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(
        (
            index
            for index, row in enumerate(rows)
            if any(_header_key(cell) == "日期" for cell in row)
        ),
        None,
    )
    if header_index is None:
        return []

    header_row = rows[header_index]
    next_row = rows[header_index + 1] if header_index + 1 < len(rows) else ()
    has_quantity_subheaders = any(
        _header_key(cell) in {"匹数", "米数"} for cell in next_row
    )
    headers: list[str] = []
    quantity_group: str | None = None
    for index, cell in enumerate(header_row):
        header = _header_key(cell)
        if header in {"入库", "出库", "库存"}:
            quantity_group = header
        elif header:
            quantity_group = None
        subheader = _header_key(next_row[index]) if has_quantity_subheaders else ""
        headers.append(
            f"{quantity_group}{subheader}"
            if quantity_group and subheader in {"匹数", "米数"}
            else header
        )
    data_start = header_index + 2 if has_quantity_subheaders else header_index + 1
    return [
        (
            number,
            {headers[index]: cell for index, cell in enumerate(row) if headers[index]},
        )
        for number, row in enumerate(rows[data_start:], start=data_start + 1)
        if any(cell is not None for cell in row)
    ]


def _value(cells: dict[str, object], *names: str) -> object:
    for name in names:
        if name in cells:
            return cells[name]
    return None


def _source_balance_snapshot(
    cells: dict[str, object], kind: LegacyWorkbookKind | str
) -> dict[str, object]:
    if LegacyWorkbookKind(kind) is LegacyWorkbookKind.RAW:
        return {"rolls": _value(cells, "库存")}
    return {
        "rolls": _value(cells, "库存匹数"),
        "meters": _value(cells, "库存米数"),
    }


def _movement_definitions(
    *,
    cells: dict[str, object],
    kind: LegacyWorkbookKind | str,
    worksheet_name: str,
) -> tuple[list[tuple[InventoryDocumentType, Decimal, Decimal | None]], str]:
    workbook_kind = LegacyWorkbookKind(kind)
    source_unit = _text(_value(cells, "加工单位"), worksheet_name) or worksheet_name
    if workbook_kind is LegacyWorkbookKind.RAW:
        rolls = _number(_value(cells, "入库", "入库匹数", "入库数量"))
        is_return = source_unit.endswith("退走") or rolls < 0
        processing_unit_name = (
            source_unit.removesuffix("退走") if is_return else source_unit
        )
        return (
            [
                (
                    InventoryDocumentType.RAW_RETURN
                    if is_return
                    else InventoryDocumentType.RAW_RECEIPT,
                    abs(rolls),
                    None,
                )
            ]
            if rolls
            else [],
            processing_unit_name,
        )

    movements: list[tuple[InventoryDocumentType, Decimal, Decimal | None]] = []
    for document_type, rolls_name, meters_name in (
        (InventoryDocumentType.FINISHED_RECEIPT, "入库匹数", "入库米数"),
        (InventoryDocumentType.FINISHED_SHIPMENT, "出库匹数", "出库米数"),
    ):
        rolls = _number(_value(cells, rolls_name))
        meters = _number(_value(cells, meters_name))
        if rolls or meters:
            normalized_type = document_type
            if rolls < 0:
                normalized_type = (
                    InventoryDocumentType.FINISHED_SHIPMENT
                    if document_type is InventoryDocumentType.FINISHED_RECEIPT
                    else InventoryDocumentType.FINISHED_RECEIPT
                )
            movements.append((normalized_type, abs(rolls), abs(meters)))
    return movements, source_unit


def import_legacy_workbooks(
    *,
    session: Session,
    raw_content: bytes,
    raw_filename: str,
    finished_content: bytes,
    finished_filename: str,
) -> LegacyInventoryExcelImportPublic:
    raw_hash, finished_hash = _sha256(raw_content), _sha256(finished_content)
    fingerprint = hashlib.sha256(
        f"{raw_hash}:{finished_hash}:{IMPORTER_VERSION}".encode()
    ).hexdigest()
    if session.exec(
        select(InventoryImportBatch).where(
            InventoryImportBatch.source_fingerprint == fingerprint
        )
    ).first():
        raise ConflictError("These workbooks were already imported")
    batch = InventoryImportBatch(
        source_fingerprint=fingerprint,
        raw_workbook_sha256=raw_hash,
        finished_workbook_sha256=finished_hash,
        importer_version=IMPORTER_VERSION,
        reconciliation_report={},
    )
    session.add(batch)
    session.flush()
    report: dict[str, int] = {
        "source_rows": 0,
        "ledger_entries": 0,
        "requires_cleanup": 0,
        "reconciliation_openings": 0,
    }
    issues: list[ExcelIssue] = []
    balances: dict[tuple[object, ...], tuple[Decimal, Decimal]] = {}
    raw_workbook = load_xlsx_workbook(raw_content)
    finished_workbook: Workbook | None = None
    try:
        finished_workbook = load_xlsx_workbook(finished_content)
        _import_book(
            session=session,
            batch=batch,
            workbook=raw_workbook,
            workbook_name=raw_filename,
            kind=LegacyWorkbookKind.RAW,
            report=report,
            balances=balances,
            issues=issues,
        )
        _import_book(
            session=session,
            batch=batch,
            workbook=finished_workbook,
            workbook_name=finished_filename,
            kind=LegacyWorkbookKind.FINISHED,
            report=report,
            balances=balances,
            issues=issues,
        )
    finally:
        raw_workbook.close()
        if finished_workbook:
            finished_workbook.close()
    if issues:
        raise ExcelValidationError(issues)
    batch.reconciliation_report = cast(dict[str, object], report)
    session.add(batch)
    session.flush()
    return LegacyInventoryExcelImportPublic(import_batch_id=batch.id, report=report)


def import_workbooks(
    *,
    session: Session,
    actor_user_id: uuid.UUID,
    raw_workbook: Path,
    finished_workbook: Path,
    dry_run: bool = False,
) -> dict[str, int]:
    actor = session.get(User, actor_user_id)
    if actor is None:
        raise BadRequestError("Import actor does not exist")
    if not actor.is_system_actor and not actor.is_active:
        raise BadRequestError(
            "Import actor must be active or a provisioned System Actor"
        )
    previous_actor_id = session.info.get(AUDIT_ACTOR_SESSION_KEY)
    bind_audit_actor(session=session, actor_id=actor_user_id)
    try:
        result = import_legacy_workbooks(
            session=session,
            raw_content=raw_workbook.read_bytes(),
            raw_filename=raw_workbook.name,
            finished_content=finished_workbook.read_bytes(),
            finished_filename=finished_workbook.name,
        )
        if dry_run:
            session.rollback()
        else:
            session.commit()
        return result.report
    except ExcelValidationError as error:
        session.rollback()
        raise BadRequestError(error.issues[0].message) from error
    except Exception:
        session.rollback()
        raise
    finally:
        if isinstance(previous_actor_id, uuid.UUID):
            bind_audit_actor(session=session, actor_id=previous_actor_id)
        else:
            clear_audit_actor(session=session)


def _has_legacy_header(sheet: Worksheet) -> bool:
    return any(
        _header_key(cell) == "日期"
        for row in sheet.iter_rows(values_only=True)
        for cell in row
    )


def _has_values(sheet: Worksheet) -> bool:
    return any(
        value is not None and str(value).strip()
        for row in sheet.iter_rows(values_only=True)
        for value in row
    )


def _import_book(
    *,
    session: Session,
    batch: InventoryImportBatch,
    workbook: Workbook,
    workbook_name: str,
    kind: LegacyWorkbookKind,
    report: dict[str, int],
    balances: dict[tuple[object, ...], tuple[Decimal, Decimal]],
    issues: list[ExcelIssue],
) -> None:
    data_rows = 0
    for sheet in workbook.worksheets:
        if not _has_legacy_header(sheet):
            if not _has_values(sheet):
                continue
            issues.append(
                ExcelIssue(
                    worksheet=sheet.title,
                    row=None,
                    column=None,
                    field=None,
                    message="Legacy worksheet does not contain a 日期 header",
                )
            )
            continue
        source_rows = _row_values(sheet)
        data_rows += len(source_rows)
        if data_rows > MAX_XLSX_DATA_ROWS:
            issues.append(
                ExcelIssue(
                    worksheet=sheet.title,
                    row=None,
                    column=None,
                    field=None,
                    message=(
                        f"Workbook exceeds the {MAX_XLSX_DATA_ROWS} data row limit"
                    ),
                )
            )
            return
        for source_row, cells in source_rows:
            report_before = report.copy()
            balances_before = balances.copy()
            try:
                with session.begin_nested():
                    _import_legacy_row(
                        session=session,
                        batch=batch,
                        workbook=workbook,
                        workbook_name=workbook_name,
                        kind=kind,
                        sheet=sheet,
                        source_row=source_row,
                        cells=cells,
                        report=report,
                        balances=balances,
                    )
            except BadRequestError as error:
                report.clear()
                report.update(report_before)
                balances.clear()
                balances.update(balances_before)
                issues.append(
                    ExcelIssue(
                        worksheet=sheet.title,
                        row=source_row,
                        column=None,
                        field=None,
                        message=str(error),
                    )
                )


def _import_legacy_row(
    *,
    session: Session,
    batch: InventoryImportBatch,
    workbook: Workbook,
    workbook_name: str,
    kind: LegacyWorkbookKind,
    sheet: Worksheet,
    source_row: int,
    cells: dict[str, object],
    report: dict[str, int],
    balances: dict[tuple[object, ...], tuple[Decimal, Decimal]],
) -> None:
    item_name = _text(_value(cells, "品名", "名称"))
    if not item_name:
        return
    movements, processing_unit_name = _movement_definitions(
        cells=cells, kind=kind, worksheet_name=sheet.title
    )
    unit = cast(
        ProcessingUnit,
        _find_or_create_unit(session, ProcessingUnit, processing_unit_name),
    )
    item_code = _text(_value(cells, "品号", "货号"), MISSING_ITEM_CODE)
    wool = (
        _text(_value(cells, "含毛量", "含毛"), MISSING_WOOL_CONTENT)
        or MISSING_WOOL_CONTENT
    )
    color = _text(_value(cells, "颜色+色号", "颜色", "色号"))
    lot = _text(
        _value(cells, "缸号"),
        MISSING_DYE_LOT if kind is LegacyWorkbookKind.FINISHED else None,
    )
    cleanup = (
        item_code == MISSING_ITEM_CODE
        or wool == MISSING_WOOL_CONTENT
        or lot == MISSING_DYE_LOT
    )
    source = LegacyImportRow(
        import_batch_id=batch.id,
        workbook_kind=kind,
        workbook_name=workbook_name,
        worksheet_name=sheet.title,
        source_row_number=source_row,
        raw_cells=json.loads(json.dumps(cells, default=str)),
        source_balance_snapshot=_source_balance_snapshot(cells, kind),
        requires_cleanup=cleanup,
    )
    session.add(source)
    session.flush()
    if cleanup:
        report["requires_cleanup"] += 1
    if movements:
        business_date = _date(_value(cells, "日期", "时间"), epoch=workbook.epoch)
        for document_type, rolls, meters in movements:
            _write_legacy_movement(
                session=session,
                batch=batch,
                source=source,
                unit=unit,
                document_type=document_type,
                item_name=item_name,
                item_code=item_code,
                wool=wool,
                color=color,
                lot=lot,
                rolls=rolls,
                meters=meters,
                business_date=business_date,
                number=_text(_value(cells, "单号", "出库单号")),
                receiving_name=_text(_value(cells, "收货单位")),
                report=report,
                balances=balances,
            )
    report["source_rows"] += 1


def _find_or_create_unit(
    session: Session,
    model: type[ProcessingUnit] | type[ReceivingUnit],
    name: str,
) -> ProcessingUnit | ReceivingUnit:
    normalized = " ".join(name.split())
    unit = session.exec(
        select(model).where(model.normalized_name == normalized)
    ).first()
    if unit:
        return cast(ProcessingUnit | ReceivingUnit, unit)
    unit = model(
        name=normalized,
        normalized_name=normalized,
    )
    session.add(unit)
    session.flush()
    return unit


def _write_legacy_movement(
    *,
    session: Session,
    batch: InventoryImportBatch,
    source: LegacyImportRow,
    unit: ProcessingUnit,
    document_type: InventoryDocumentType,
    item_name: str,
    item_code: str | None,
    wool: str,
    color: str | None,
    lot: str | None,
    rolls: Decimal,
    meters: Decimal | None,
    business_date: date,
    number: str | None,
    receiving_name: str | None,
    report: dict[str, int],
    balances: dict[tuple[object, ...], tuple[Decimal, Decimal]],
) -> None:
    is_finished = document_type in {
        InventoryDocumentType.FINISHED_RECEIPT,
        InventoryDocumentType.FINISHED_SHIPMENT,
    }
    direction = (
        -1
        if document_type
        in {
            InventoryDocumentType.RAW_RETURN,
            InventoryDocumentType.FINISHED_SHIPMENT,
        }
        else 1
    )
    receiving = (
        _find_or_create_unit(
            session,
            ReceivingUnit,
            receiving_name or "历史未填写收货单位",
        )
        if document_type is InventoryDocumentType.FINISHED_SHIPMENT
        else None
    )
    document = InventoryDocument(
        document_type=document_type,
        business_date=business_date,
        processing_unit_id=unit.id,
        receiving_unit_id=receiving.id if receiving else None,
        document_number=number,
        is_legacy=True,
    )
    session.add(document)
    session.flush()
    line = InventoryDocumentLine(
        document_id=document.id,
        line_no=1,
        item_name=item_name,
        item_code=item_code if not is_finished else None,
        wool_content=wool,
        color_code=color if is_finished else None,
        dye_lot_no=lot if is_finished else None,
        quantity_rolls=rolls,
        quantity_meters=meters if is_finished and meters else None,
    )
    session.add(line)
    session.flush()
    ledger_kind = (
        InventoryLedgerKind.FINISHED if is_finished else InventoryLedgerKind.RAW
    )
    ledger_item_code = item_code if not is_finished else None
    ledger_color = color if is_finished else None
    ledger_lot = lot if is_finished else None
    meter_delta = direction * (meters or Decimal("0"))
    rolls_delta = direction * rolls
    balance_key = (
        ledger_kind,
        unit.id,
        item_name,
        ledger_item_code,
        wool,
        ledger_color,
        ledger_lot,
    )
    balance_rolls, balance_meters = balances.get(
        balance_key, (Decimal("0"), Decimal("0"))
    )
    opening_rolls = max(Decimal("0"), -(balance_rolls + rolls_delta))
    opening_meters = max(Decimal("0"), -(balance_meters + meter_delta))
    if opening_rolls or opening_meters:
        session.add(
            InventoryLedgerEntry(
                ledger_kind=ledger_kind,
                movement_type=InventoryMovementType.MIGRATION_RECONCILIATION_OPENING,
                business_date=business_date,
                processing_unit_id=unit.id,
                legacy_import_row_id=source.id,
                import_batch_id=batch.id,
                item_name=item_name,
                item_code=ledger_item_code,
                wool_content=wool,
                color_code=ledger_color,
                dye_lot_no=ledger_lot,
                rolls_delta=opening_rolls,
                meters_delta=opening_meters,
                reason="历史迁移对账期初",
            )
        )
        report["ledger_entries"] += 1
        report["reconciliation_openings"] += 1
        balance_rolls += opening_rolls
        balance_meters += opening_meters
    session.add(
        InventoryLedgerEntry(
            ledger_kind=ledger_kind,
            movement_type=InventoryMovementType(document_type),
            business_date=business_date,
            processing_unit_id=unit.id,
            document_line_id=line.id,
            legacy_import_row_id=source.id,
            import_batch_id=batch.id,
            item_name=item_name,
            item_code=ledger_item_code,
            wool_content=wool,
            color_code=ledger_color,
            dye_lot_no=ledger_lot,
            rolls_delta=rolls_delta,
            meters_delta=meter_delta,
        )
    )
    balances[balance_key] = (balance_rolls + rolls_delta, balance_meters + meter_delta)
    report["ledger_entries"] += 1
