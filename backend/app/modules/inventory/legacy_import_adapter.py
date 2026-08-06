import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import cast

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.utils.datetime import from_excel  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from app.core.excel import (
    MAX_XLSX_DATA_ROWS,
    ExcelIssue,
    ExcelValidationError,
    load_xlsx_workbook,
)
from app.core.exceptions import BadRequestError
from app.models.inventory import (
    InventoryDocumentType,
    LegacyWorkbookKind,
)

MISSING_ITEM_CODE = "未填写品号"
MISSING_WOOL_CONTENT = "未填写含毛量"
MISSING_DYE_LOT = "未分缸"


@dataclass(frozen=True)
class LegacyMovementRecord:
    document_type: InventoryDocumentType
    rolls: Decimal
    meters: Decimal | None
    business_date: date
    document_number: str | None
    receiving_unit_name: str | None


@dataclass(frozen=True)
class LegacyImportRecord:
    workbook_kind: LegacyWorkbookKind
    workbook_name: str
    worksheet_name: str
    source_row_number: int
    raw_cells: dict[str, object]
    source_balance_snapshot: dict[str, object]
    requires_cleanup: bool
    item_name: str
    processing_unit_name: str
    item_code: str | None
    wool_content: str
    color_code: str | None
    dye_lot_no: str | None
    movements: tuple[LegacyMovementRecord, ...]


def read_legacy_workbooks(
    *,
    raw_content: bytes,
    raw_filename: str,
    finished_content: bytes,
    finished_filename: str,
) -> tuple[list[LegacyImportRecord], list[ExcelIssue]]:
    try:
        raw_workbook = load_xlsx_workbook(raw_content)
    except ExcelValidationError as error:
        return [], error.issues

    finished_workbook: Workbook | None = None
    records: list[LegacyImportRecord] = []
    issues: list[ExcelIssue] = []
    try:
        try:
            finished_workbook = load_xlsx_workbook(finished_content)
        except ExcelValidationError as error:
            return [], error.issues
        _read_book(
            workbook=raw_workbook,
            workbook_name=raw_filename,
            kind=LegacyWorkbookKind.RAW,
            records=records,
            issues=issues,
        )
        _read_book(
            workbook=finished_workbook,
            workbook_name=finished_filename,
            kind=LegacyWorkbookKind.FINISHED,
            records=records,
            issues=issues,
        )
    finally:
        raw_workbook.close()
        if finished_workbook is not None:
            finished_workbook.close()
    return records, issues


def _text(value: object, placeholder: str | None = None) -> str | None:
    if value is None or not str(value).strip():
        return placeholder
    return str(value).strip()


def _number(value: object) -> Decimal:
    try:
        return Decimal(str(value or 0).replace(",", ""))
    except Exception as err:
        raise BadRequestError(f"Invalid inventory quantity: {value}") from err


def _date(value: object, *, epoch: datetime) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value or "").strip()
    year_closing_match = re.fullmatch(r"(\d{4})年结存", text)
    if year_closing_match:
        return date(int(year_closing_match.group(1)), 12, 31)

    try:
        numeric_value = Decimal(text)
    except Exception:
        numeric_value = None
    if numeric_value is not None and 30_000 <= numeric_value <= 100_000:
        converted = from_excel(float(numeric_value), epoch)
        return converted.date() if isinstance(converted, datetime) else converted

    slash_date_match = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", text)
    if slash_date_match:
        return date(
            int(slash_date_match.group(1)),
            int(slash_date_match.group(2)),
            int(slash_date_match.group(3)),
        )

    try:
        return datetime.fromisoformat(text.replace("/", "-")).date()
    except ValueError as err:
        raise BadRequestError(f"Invalid inventory date: {value}") from err


def _header_key(value: object) -> str:
    return "".join(str(value or "").strip().lower().split())


def _row_values(sheet: Worksheet) -> list[tuple[int, dict[str, object]]]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(
        (
            index
            for index, row in enumerate(rows)
            if any(_header_key(cell) == "日期" for cell in row)
        ),
        None,
    )
    if header_index is None:
        return []

    header_row = rows[header_index]
    next_row = rows[header_index + 1] if header_index + 1 < len(rows) else ()
    has_quantity_subheaders = any(
        _header_key(cell) in {"匹数", "米数"} for cell in next_row
    )
    headers: list[str] = []
    quantity_group: str | None = None
    for index, cell in enumerate(header_row):
        header = _header_key(cell)
        if header in {"入库", "出库", "库存"}:
            quantity_group = header
        elif header:
            quantity_group = None
        subheader = _header_key(next_row[index]) if has_quantity_subheaders else ""
        headers.append(
            f"{quantity_group}{subheader}"
            if quantity_group and subheader in {"匹数", "米数"}
            else header
        )
    data_start = header_index + 2 if has_quantity_subheaders else header_index + 1
    return [
        (
            number,
            {headers[index]: cell for index, cell in enumerate(row) if headers[index]},
        )
        for number, row in enumerate(rows[data_start:], start=data_start + 1)
        if any(cell is not None for cell in row)
    ]


def _value(cells: dict[str, object], *names: str) -> object:
    for name in names:
        if name in cells:
            return cells[name]
    return None


def _source_balance_snapshot(
    cells: dict[str, object], kind: LegacyWorkbookKind | str
) -> dict[str, object]:
    if LegacyWorkbookKind(kind) is LegacyWorkbookKind.RAW:
        return {"rolls": _value(cells, "库存")}
    return {
        "rolls": _value(cells, "库存匹数"),
        "meters": _value(cells, "库存米数"),
    }


def _movement_definitions(
    *,
    cells: dict[str, object],
    kind: LegacyWorkbookKind | str,
    worksheet_name: str,
) -> tuple[list[tuple[InventoryDocumentType, Decimal, Decimal | None]], str]:
    workbook_kind = LegacyWorkbookKind(kind)
    source_unit = _text(_value(cells, "加工单位"), worksheet_name) or worksheet_name
    if workbook_kind is LegacyWorkbookKind.RAW:
        rolls = _number(_value(cells, "入库", "入库匹数", "入库数量"))
        is_return = source_unit.endswith("退走") or rolls < 0
        processing_unit_name = (
            source_unit.removesuffix("退走") if is_return else source_unit
        )
        return (
            [
                (
                    InventoryDocumentType.RAW_RETURN
                    if is_return
                    else InventoryDocumentType.RAW_RECEIPT,
                    abs(rolls),
                    None,
                )
            ]
            if rolls
            else [],
            processing_unit_name,
        )

    movements: list[tuple[InventoryDocumentType, Decimal, Decimal | None]] = []
    for document_type, rolls_name, meters_name in (
        (InventoryDocumentType.FINISHED_RECEIPT, "入库匹数", "入库米数"),
        (InventoryDocumentType.FINISHED_SHIPMENT, "出库匹数", "出库米数"),
    ):
        rolls = _number(_value(cells, rolls_name))
        meters = _number(_value(cells, meters_name))
        if rolls or meters:
            normalized_type = document_type
            if rolls < 0:
                normalized_type = (
                    InventoryDocumentType.FINISHED_SHIPMENT
                    if document_type is InventoryDocumentType.FINISHED_RECEIPT
                    else InventoryDocumentType.FINISHED_RECEIPT
                )
            movements.append((normalized_type, abs(rolls), abs(meters)))
    return movements, source_unit


def _has_legacy_header(sheet: Worksheet) -> bool:
    return any(
        _header_key(cell) == "日期"
        for row in sheet.iter_rows(values_only=True)
        for cell in row
    )


def _has_values(sheet: Worksheet) -> bool:
    return any(
        value is not None and str(value).strip()
        for row in sheet.iter_rows(values_only=True)
        for value in row
    )


def _read_book(
    *,
    workbook: Workbook,
    workbook_name: str,
    kind: LegacyWorkbookKind,
    records: list[LegacyImportRecord],
    issues: list[ExcelIssue],
) -> None:
    data_rows = 0
    for sheet in workbook.worksheets:
        if not _has_legacy_header(sheet):
            if not _has_values(sheet):
                continue
            issues.append(
                ExcelIssue(
                    worksheet=sheet.title,
                    row=None,
                    column=None,
                    field=None,
                    message="Legacy worksheet does not contain a 日期 header",
                )
            )
            continue
        source_rows = _row_values(sheet)
        data_rows += len(source_rows)
        if data_rows > MAX_XLSX_DATA_ROWS:
            issues.append(
                ExcelIssue(
                    worksheet=sheet.title,
                    row=None,
                    column=None,
                    field=None,
                    message=f"Workbook exceeds the {MAX_XLSX_DATA_ROWS} data row limit",
                )
            )
            return
        for source_row, cells in source_rows:
            try:
                record = _normalize_row(
                    workbook=workbook,
                    workbook_name=workbook_name,
                    kind=kind,
                    sheet=sheet,
                    source_row=source_row,
                    cells=cells,
                )
            except BadRequestError as error:
                issues.append(
                    ExcelIssue(
                        worksheet=sheet.title,
                        row=source_row,
                        column=None,
                        field=None,
                        message=str(error),
                    )
                )
                continue
            if record is not None:
                records.append(record)


def _normalize_row(
    *,
    workbook: Workbook,
    workbook_name: str,
    kind: LegacyWorkbookKind,
    sheet: Worksheet,
    source_row: int,
    cells: dict[str, object],
) -> LegacyImportRecord | None:
    item_name = _text(_value(cells, "品名", "名称"))
    if not item_name:
        return None
    movements, processing_unit_name = _movement_definitions(
        cells=cells, kind=kind, worksheet_name=sheet.title
    )
    item_code = _text(_value(cells, "品号", "货号"), MISSING_ITEM_CODE)
    wool = (
        _text(_value(cells, "含毛量", "含毛"), MISSING_WOOL_CONTENT)
        or MISSING_WOOL_CONTENT
    )
    color = _text(_value(cells, "颜色+色号", "颜色", "色号"))
    lot = _text(
        _value(cells, "缸号"),
        MISSING_DYE_LOT if kind is LegacyWorkbookKind.FINISHED else None,
    )
    cleanup = (
        item_code == MISSING_ITEM_CODE
        or wool == MISSING_WOOL_CONTENT
        or lot == MISSING_DYE_LOT
    )
    business_date = None
    if movements:
        business_date = _date(_value(cells, "日期", "时间"), epoch=workbook.epoch)
    return LegacyImportRecord(
        workbook_kind=kind,
        workbook_name=workbook_name,
        worksheet_name=sheet.title,
        source_row_number=source_row,
        raw_cells=cast(dict[str, object], json.loads(json.dumps(cells, default=str))),
        source_balance_snapshot=_source_balance_snapshot(cells, kind),
        requires_cleanup=cleanup,
        item_name=item_name,
        processing_unit_name=processing_unit_name,
        item_code=item_code,
        wool_content=wool,
        color_code=color,
        dye_lot_no=lot,
        movements=tuple(
            LegacyMovementRecord(
                document_type=document_type,
                rolls=rolls,
                meters=meters,
                business_date=cast(date, business_date),
                document_number=_text(_value(cells, "单号", "出库单号")),
                receiving_unit_name=_text(_value(cells, "收货单位")),
            )
            for document_type, rolls, meters in movements
        ),
    )
