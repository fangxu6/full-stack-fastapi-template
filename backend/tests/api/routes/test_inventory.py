import inspect
import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlmodel import Session

from app.core.config import settings
from app.models import InventoryDocument
from app.models.base import get_datetime_utc
from app.modules.inventory.router import (
    import_documents_from_excel,
    import_legacy_workbooks_from_excel,
)
from app.schemas.inventory import InventoryLinePublic

INVENTORY_PATH = f"{settings.API_V1_STR}/inventory"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_excel_import_handlers_are_synchronous() -> None:
    assert not inspect.iscoroutinefunction(import_documents_from_excel)
    assert not inspect.iscoroutinefunction(import_legacy_workbooks_from_excel)


def _decimal(value: object) -> Decimal:
    return Decimal(str(value))


def _create_processing_unit(
    client: TestClient, headers: dict[str, str]
) -> dict[str, str]:
    response = client.post(
        f"{INVENTORY_PATH}/processing-units",
        headers=headers,
        json={"name": f"Unit {uuid.uuid4()}"},
    )
    assert response.status_code == 200
    return response.json()


def _create_receiving_unit(
    client: TestClient, headers: dict[str, str]
) -> dict[str, str]:
    response = client.post(
        f"{INVENTORY_PATH}/receiving-units",
        headers=headers,
        json={"name": f"Customer {uuid.uuid4()}"},
    )
    assert response.status_code == 200
    return response.json()


def _create_raw_receipt(
    client: TestClient,
    headers: dict[str, str],
    *,
    processing_unit_id: str,
    item_name: str,
    item_code: str,
    document_number: str | None = None,
    business_date: date | None = None,
    quantity_rolls: str = "1",
) -> dict[str, object]:
    payload = {
        "document_type": "RAW_RECEIPT",
        "business_date": str(business_date or date.today()),
        "processing_unit_id": processing_unit_id,
        "document_number": document_number or f"R-{uuid.uuid4()}",
        "lines": [
            {
                "item_name": item_name,
                "item_code": item_code,
                "wool_content": "100% wool",
                "quantity_rolls": quantity_rolls,
            }
        ],
    }
    response = client.post(f"{INVENTORY_PATH}/documents", headers=headers, json=payload)
    assert response.status_code == 200, response.json()
    return response.json()


def _xlsx_bytes(rows: list[list[object]], *, sheet_name: str = "单据导入") -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.parametrize(
    ("endpoint", "factory"),
    [
        ("processing-units", _create_processing_unit),
        ("receiving-units", _create_receiving_unit),
    ],
)
def test_master_unit_lists_are_offset_paginated(
    client: TestClient,
    superuser_token_headers: dict[str, str],
    endpoint: str,
    factory: object,
) -> None:
    for _ in range(3):
        factory(client, superuser_token_headers)

    response = client.get(
        f"{INVENTORY_PATH}/{endpoint}",
        headers=superuser_token_headers,
        params={"skip": 1, "limit": 2},
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload["data"]) == 2
    assert payload["count"] >= 3
    assert payload["count"] > len(payload["data"])


@pytest.mark.parametrize("endpoint", ["processing-units", "receiving-units"])
def test_master_unit_lists_filter_by_name_and_status(
    client: TestClient,
    superuser_token_headers: dict[str, str],
    endpoint: str,
) -> None:
    prefix = f"Filter {uuid.uuid4()}"
    endpoint_url = f"{INVENTORY_PATH}/{endpoint}"
    active_response = client.post(
        endpoint_url,
        headers=superuser_token_headers,
        json={"name": f"{prefix} Active"},
    )
    inactive_response = client.post(
        endpoint_url,
        headers=superuser_token_headers,
        json={"name": f"{prefix} Inactive"},
    )
    unrelated_response = client.post(
        endpoint_url,
        headers=superuser_token_headers,
        json={"name": f"Other {uuid.uuid4()}"},
    )
    assert active_response.status_code == 200
    assert inactive_response.status_code == 200
    assert unrelated_response.status_code == 200

    update_response = client.put(
        f"{endpoint_url}/{inactive_response.json()['id']}",
        headers=superuser_token_headers,
        json={"is_active": False},
    )
    assert update_response.status_code == 200

    filtered_response = client.get(
        endpoint_url,
        headers=superuser_token_headers,
        params={"name": f"  {prefix}  ", "skip": 0, "limit": 20},
    )
    assert filtered_response.status_code == 200
    filtered_payload = filtered_response.json()
    assert filtered_payload["count"] == 2
    assert {unit["name"] for unit in filtered_payload["data"]} == {
        f"{prefix} Active",
        f"{prefix} Inactive",
    }

    active_filtered_response = client.get(
        endpoint_url,
        headers=superuser_token_headers,
        params={"name": prefix, "is_active": True},
    )
    assert active_filtered_response.status_code == 200
    assert active_filtered_response.json()["count"] == 1
    assert active_filtered_response.json()["data"][0]["name"] == f"{prefix} Active"

    inactive_filtered_response = client.get(
        endpoint_url,
        headers=superuser_token_headers,
        params={"name": prefix, "is_active": False},
    )
    assert inactive_filtered_response.status_code == 200
    assert inactive_filtered_response.json()["count"] == 1
    assert inactive_filtered_response.json()["data"][0]["name"] == f"{prefix} Inactive"


def test_document_list_paginates_after_filters(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    prefix = f"PG-{uuid.uuid4()}"
    for index in range(3):
        _create_raw_receipt(
            client,
            superuser_token_headers,
            processing_unit_id=processing_unit["id"],
            item_name=f"Raw fabric {prefix}",
            item_code=f"RF-{uuid.uuid4()}",
            document_number=f"{prefix}-{index}",
        )

    response = client.get(
        f"{INVENTORY_PATH}/documents",
        headers=superuser_token_headers,
        params={"document_number": prefix, "skip": 1, "limit": 2},
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload["data"]) == 2
    assert payload["count"] == 3


def test_balance_list_paginates_aggregated_rows(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    prefix = f"Balance {uuid.uuid4()}"
    for index in range(3):
        _create_raw_receipt(
            client,
            superuser_token_headers,
            processing_unit_id=processing_unit["id"],
            item_name=f"{prefix}-{index}",
            item_code=f"RB-{uuid.uuid4()}",
        )

    response = client.get(
        f"{INVENTORY_PATH}/balances/raw",
        headers=superuser_token_headers,
        params={"item_name": prefix, "skip": 1, "limit": 2},
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload["data"]) == 2
    assert payload["count"] == 3


def test_ledger_list_paginates_matching_entries(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    item_name = f"Ledger {uuid.uuid4()}"
    item_code = f"RL-{uuid.uuid4()}"
    for day in range(1, 4):
        _create_raw_receipt(
            client,
            superuser_token_headers,
            processing_unit_id=processing_unit["id"],
            item_name=item_name,
            item_code=item_code,
            business_date=date(2026, 7, day),
        )

    response = client.get(
        f"{INVENTORY_PATH}/ledger",
        headers=superuser_token_headers,
        params={
            "ledger_kind": "RAW",
            "processing_unit_id": processing_unit["id"],
            "item_name": item_name,
            "item_code": item_code,
            "wool_content": "100% wool",
            "skip": 1,
            "limit": 1,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload["data"]) == 1
    assert payload["count"] == 3


@pytest.mark.parametrize(
    "params",
    [
        {"skip": -1, "limit": 20},
        {"skip": 0, "limit": 0},
        {"skip": 0, "limit": 101},
    ],
)
def test_inventory_lists_reject_invalid_pagination(
    client: TestClient,
    superuser_token_headers: dict[str, str],
    params: dict[str, int],
) -> None:
    response = client.get(
        f"{INVENTORY_PATH}/documents",
        headers=superuser_token_headers,
        params=params,
    )

    assert response.status_code == 422
    assert response.json()["request_id"]


def test_raw_receipt_return_and_balance(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    item_code = f"RF-{uuid.uuid4()}"
    receipt = {
        "document_type": "RAW_RECEIPT",
        "business_date": str(date.today()),
        "processing_unit_id": processing_unit["id"],
        "document_number": f"R-{uuid.uuid4()}",
        "lines": [
            {
                "item_name": "Raw fabric",
                "item_code": item_code,
                "wool_content": "100% wool",
                "quantity_rolls": "5.50",
            }
        ],
    }
    response = client.post(
        f"{INVENTORY_PATH}/documents", headers=superuser_token_headers, json=receipt
    )
    assert response.status_code == 200

    return_document = {
        **receipt,
        "document_type": "RAW_RETURN",
        "document_number": f"T-{uuid.uuid4()}",
    }
    return_document["lines"] = [{**receipt["lines"][0], "quantity_rolls": "2.25"}]
    response = client.post(
        f"{INVENTORY_PATH}/documents",
        headers=superuser_token_headers,
        json=return_document,
    )
    assert response.status_code == 200

    response = client.get(
        f"{INVENTORY_PATH}/balances/raw", headers=superuser_token_headers
    )
    assert response.status_code == 200
    balance = next(
        item for item in response.json()["data"] if item["item_code"] == item_code
    )
    assert _decimal(balance["rolls_balance"]) == Decimal("3.25")


def test_raw_return_rejects_negative_balance(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    item_code = f"RF-{uuid.uuid4()}"
    payload = {
        "document_type": "RAW_RETURN",
        "business_date": str(date.today()),
        "processing_unit_id": processing_unit["id"],
        "document_number": f"T-{uuid.uuid4()}",
        "lines": [
            {
                "item_name": "Raw fabric",
                "item_code": item_code,
                "wool_content": "100% wool",
                "quantity_rolls": 1,
            }
        ],
    }
    response = client.post(
        f"{INVENTORY_PATH}/documents", headers=superuser_token_headers, json=payload
    )
    assert response.status_code == 409
    assert response.json()["detail"] == "Insufficient inventory"
    assert response.json()["request_id"]


def test_ledger_affected_documents_require_correction(
    client: TestClient, superuser_token_headers: dict[str, str], db: Session
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    item_code = f"RF-{uuid.uuid4()}"
    document = {
        "document_type": "RAW_RECEIPT",
        "business_date": str(date.today()),
        "processing_unit_id": processing_unit["id"],
        "document_number": f"R-{uuid.uuid4()}",
        "lines": [
            {
                "item_name": "Raw fabric",
                "item_code": item_code,
                "wool_content": "100% wool",
                "quantity_rolls": "5.50",
            }
        ],
    }
    created = client.post(
        f"{INVENTORY_PATH}/documents",
        headers=superuser_token_headers,
        json=document,
    )
    assert created.status_code == 200
    document_id = created.json()["id"]

    document["lines"][0]["quantity_rolls"] = "3.25"
    updated = client.put(
        f"{INVENTORY_PATH}/documents/{document_id}",
        headers=superuser_token_headers,
        json=document,
    )
    assert updated.status_code == 409
    assert updated.json()["detail"] == "INVENTORY_CORRECTION_REQUIRED"
    assert updated.json()["request_id"]

    deleted = client.delete(
        f"{INVENTORY_PATH}/documents/{document_id}", headers=superuser_token_headers
    )
    assert deleted.status_code == 409
    assert deleted.json()["detail"] == "INVENTORY_CORRECTION_REQUIRED"
    assert deleted.json()["request_id"]

    stored_document = db.get(InventoryDocument, uuid.UUID(document_id))
    assert stored_document is not None
    stored_document.deleted_at = get_datetime_utc()
    db.add(stored_document)
    db.commit()

    restored = client.post(
        f"{INVENTORY_PATH}/documents/{document_id}/restore",
        headers=superuser_token_headers,
    )
    assert restored.status_code == 409
    assert restored.json()["detail"] == "INVENTORY_CORRECTION_REQUIRED"
    assert restored.json()["request_id"]


def test_finished_shipment_updates_roll_and_meter_balances(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    receiving = client.post(
        f"{INVENTORY_PATH}/receiving-units",
        headers=superuser_token_headers,
        json={"name": f"Customer {uuid.uuid4()}"},
    )
    assert receiving.status_code == 200
    shipment = {
        "document_type": "FINISHED_SHIPMENT",
        "business_date": str(date.today()),
        "processing_unit_id": processing_unit["id"],
        "receiving_unit_id": receiving.json()["id"],
        "document_number": f"S-{uuid.uuid4()}",
        "lines": [
            {
                "item_name": "Finished fabric",
                "wool_content": "70% wool",
                "color_code": "Blue-01",
                "dye_lot_no": "LOT-01",
                "quantity_rolls": "0.50",
                "quantity_meters": "12.5",
            }
        ],
    }

    response = client.post(
        f"{INVENTORY_PATH}/documents", headers=superuser_token_headers, json=shipment
    )

    assert response.status_code == 409
    assert response.json()["detail"] == "Insufficient inventory"


@pytest.mark.parametrize("quantity_rolls", ["0", "-0.01", "1.001"])
def test_roll_quantity_rejects_non_positive_or_over_precision_values(
    client: TestClient,
    superuser_token_headers: dict[str, str],
    quantity_rolls: str,
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    payload = {
        "document_type": "RAW_RECEIPT",
        "business_date": str(date.today()),
        "processing_unit_id": processing_unit["id"],
        "document_number": f"R-{uuid.uuid4()}",
        "lines": [
            {
                "item_name": "Raw fabric",
                "item_code": f"RF-{uuid.uuid4()}",
                "wool_content": "100% wool",
                "quantity_rolls": quantity_rolls,
            }
        ],
    }

    response = client.post(
        f"{INVENTORY_PATH}/documents", headers=superuser_token_headers, json=payload
    )

    assert response.status_code == 422
    assert response.json()["request_id"]


def test_historical_read_model_allows_zero_rolls() -> None:
    line = InventoryLinePublic(
        id=uuid.uuid4(),
        line_no=1,
        item_name="Legacy finished fabric",
        item_code=None,
        wool_content="70% wool",
        color_code="Blue-01",
        dye_lot_no="LOT-01",
        quantity_rolls=Decimal("0"),
        quantity_meters=Decimal("20"),
    )

    assert line.quantity_rolls == Decimal("0")


def test_inactive_processing_unit_rejects_new_document(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    deactivated = client.put(
        f"{INVENTORY_PATH}/processing-units/{processing_unit['id']}",
        headers=superuser_token_headers,
        json={"is_active": False},
    )
    assert deactivated.status_code == 200
    payload = {
        "document_type": "RAW_RECEIPT",
        "business_date": str(date.today()),
        "processing_unit_id": processing_unit["id"],
        "document_number": f"R-{uuid.uuid4()}",
        "lines": [
            {
                "item_name": "Raw fabric",
                "item_code": f"RF-{uuid.uuid4()}",
                "wool_content": "100% wool",
                "quantity_rolls": 1,
            }
        ],
    }

    response = client.post(
        f"{INVENTORY_PATH}/documents", headers=superuser_token_headers, json=payload
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "Processing unit is not active"


def test_document_number_and_legacy_placeholders_are_rejected(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    number = f"R-{uuid.uuid4()}"
    payload = {
        "document_type": "RAW_RECEIPT",
        "business_date": str(date.today()),
        "processing_unit_id": processing_unit["id"],
        "document_number": number,
        "lines": [
            {
                "item_name": "Raw fabric",
                "item_code": f"RF-{uuid.uuid4()}",
                "wool_content": "100% wool",
                "quantity_rolls": 1,
            }
        ],
    }
    assert (
        client.post(
            f"{INVENTORY_PATH}/documents", headers=superuser_token_headers, json=payload
        ).status_code
        == 200
    )

    duplicate = client.post(
        f"{INVENTORY_PATH}/documents", headers=superuser_token_headers, json=payload
    )
    assert duplicate.status_code == 409

    placeholder = client.post(
        f"{INVENTORY_PATH}/documents",
        headers=superuser_token_headers,
        json={
            **payload,
            "document_number": f"R-{uuid.uuid4()}",
            "lines": [{**payload["lines"][0], "item_code": "未填写品号"}],
        },
    )
    assert placeholder.status_code == 400
    assert (
        placeholder.json()["detail"]
        == "Legacy placeholder values cannot be used in new documents"
    )


def test_legacy_documents_cannot_be_updated_deleted_or_restored(
    client: TestClient,
    superuser_token_headers: dict[str, str],
    db: Session,
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    document = _create_raw_receipt(
        client,
        superuser_token_headers,
        processing_unit_id=processing_unit["id"],
        item_name="Legacy fabric",
        item_code=f"LEGACY-{uuid.uuid4()}",
    )
    stored_document = db.get(InventoryDocument, uuid.UUID(str(document["id"])))
    assert stored_document is not None
    stored_document.is_legacy = True
    db.add(stored_document)
    db.commit()

    update_response = client.put(
        f"{INVENTORY_PATH}/documents/{stored_document.id}",
        headers=superuser_token_headers,
        json={
            "document_type": "RAW_RECEIPT",
            "business_date": str(date.today()),
            "processing_unit_id": processing_unit["id"],
            "document_number": f"LEGACY-EDIT-{uuid.uuid4()}",
            "lines": [
                {
                    "item_name": "Legacy fabric",
                    "item_code": "LEGACY-CODE",
                    "wool_content": "100% wool",
                    "quantity_rolls": 1,
                }
            ],
        },
    )
    delete_response = client.delete(
        f"{INVENTORY_PATH}/documents/{stored_document.id}",
        headers=superuser_token_headers,
    )
    restore_response = client.post(
        f"{INVENTORY_PATH}/documents/{stored_document.id}/restore",
        headers=superuser_token_headers,
    )

    assert update_response.status_code == 400
    assert (
        update_response.json()["detail"]
        == "Legacy inventory documents cannot be edited"
    )
    assert delete_response.status_code == 400
    assert (
        delete_response.json()["detail"]
        == "Legacy inventory documents cannot be deleted"
    )
    assert restore_response.status_code == 400
    assert (
        restore_response.json()["detail"]
        == "Legacy inventory documents cannot be restored"
    )


def test_suggestions_return_saved_values(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    item_name = f"Raw {uuid.uuid4()}"
    receipt = {
        "document_type": "RAW_RECEIPT",
        "business_date": str(date.today()),
        "processing_unit_id": processing_unit["id"],
        "document_number": f"R-{uuid.uuid4()}",
        "lines": [
            {
                "item_name": item_name,
                "item_code": f"RF-{uuid.uuid4()}",
                "wool_content": "100% wool",
                "quantity_rolls": 1,
            }
        ],
    }
    assert (
        client.post(
            f"{INVENTORY_PATH}/documents", headers=superuser_token_headers, json=receipt
        ).status_code
        == 200
    )

    response = client.get(
        f"{INVENTORY_PATH}/suggestions",
        headers=superuser_token_headers,
        params={"ledger_kind": "RAW", "field": "item_name", "query": item_name},
    )

    assert response.status_code == 200
    assert response.json()["data"] == [item_name]


def test_excel_document_template_and_import_create_multiple_documents(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    template = client.get(
        f"{INVENTORY_PATH}/excel/templates/documents", headers=superuser_token_headers
    )
    assert template.status_code == 200
    assert "attachment" in template.headers["content-disposition"]
    worksheet = load_workbook(BytesIO(template.content), read_only=True)["单据导入"]
    headers = next(worksheet.values)
    assert headers == (
        "单据类型",
        "日期",
        "单据号",
        "加工单位",
        "收货单位",
        "备注",
        "品名",
        "货号",
        "含毛量",
        "颜色",
        "缸号",
        "匹数",
        "米数",
    )

    first_number = f"XLSX-{uuid.uuid4()}"
    second_number = f"XLSX-{uuid.uuid4()}"
    rows = [
        list(headers),
        [
            "RAW_RECEIPT",
            "2026-08-01",
            first_number,
            processing_unit["name"],
            None,
            "首张单据",
            "坯布 A",
            "A-001",
            "100%",
            None,
            None,
            2,
            None,
        ],
        [
            "RAW_RECEIPT",
            "2026-08-01",
            first_number,
            processing_unit["name"],
            None,
            "首张单据",
            "坯布 B",
            "B-001",
            "80%",
            None,
            None,
            1,
            None,
        ],
        [
            "RAW_RECEIPT",
            "2026-08-02",
            second_number,
            processing_unit["name"],
            None,
            None,
            "坯布 C",
            "C-001",
            "70%",
            None,
            None,
            3,
            None,
        ],
    ]
    response = client.post(
        f"{INVENTORY_PATH}/excel/imports/documents",
        headers=superuser_token_headers,
        files={
            "workbook": (
                "documents.xlsx",
                _xlsx_bytes(rows),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201, response.json()
    assert response.json() == {
        "created_documents": 2,
        "document_numbers": [first_number, second_number],
    }


def test_excel_document_import_rolls_back_the_whole_workbook(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    first_number = f"XLSX-GOOD-{uuid.uuid4()}"
    failed_number = f"XLSX-BAD-{uuid.uuid4()}"
    headers = [
        "单据类型",
        "日期",
        "单据号",
        "加工单位",
        "收货单位",
        "备注",
        "品名",
        "货号",
        "含毛量",
        "颜色",
        "缸号",
        "匹数",
        "米数",
    ]
    rows = [
        headers,
        [
            "RAW_RECEIPT",
            "2026-08-01",
            first_number,
            processing_unit["name"],
            None,
            None,
            "回滚坯布",
            "ROLLBACK-001",
            "100%",
            None,
            None,
            1,
            None,
        ],
        [
            "RAW_RETURN",
            "2026-08-02",
            failed_number,
            processing_unit["name"],
            None,
            None,
            "回滚坯布",
            "ROLLBACK-001",
            "100%",
            None,
            None,
            2,
            None,
        ],
    ]
    response = client.post(
        f"{INVENTORY_PATH}/excel/imports/documents",
        headers=superuser_token_headers,
        files={"workbook": ("rollback.xlsx", _xlsx_bytes(rows), XLSX_MEDIA_TYPE)},
    )

    assert response.status_code == 422
    assert response.json()["detail"]["issues"][0]["column"] == "单据号"
    assert response.json()["request_id"]
    documents = client.get(
        f"{INVENTORY_PATH}/documents",
        headers=superuser_token_headers,
        params={"document_number": "XLSX-GOOD"},
    )
    assert documents.status_code == 200
    assert documents.json()["count"] == 0


def test_excel_document_import_rejects_document_types_outside_page_scope(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    raw_number = f"XLSX-RAW-{uuid.uuid4()}"
    rows = [
        [
            "单据类型",
            "日期",
            "单据号",
            "加工单位",
            "收货单位",
            "备注",
            "品名",
            "货号",
            "含毛量",
            "颜色",
            "缸号",
            "匹数",
            "米数",
        ],
        [
            "RAW_RECEIPT",
            "2026-08-01",
            raw_number,
            processing_unit["name"],
            None,
            None,
            "可回滚坯布",
            "SCOPE-RAW",
            "100%",
            None,
            None,
            1,
            None,
        ],
        [
            "FINISHED_SHIPMENT",
            "2026-08-01",
            f"XLSX-FINISHED-{uuid.uuid4()}",
            processing_unit["name"],
            "无需解析",
            None,
            "越界成品",
            None,
            "70%",
            "蓝",
            "LOT-SCOPE",
            1,
            1,
        ],
    ]

    response = client.post(
        f"{INVENTORY_PATH}/excel/imports/documents",
        headers=superuser_token_headers,
        params=[("document_types", "RAW_RECEIPT"), ("document_types", "RAW_RETURN")],
        files={"workbook": ("scope.xlsx", _xlsx_bytes(rows), XLSX_MEDIA_TYPE)},
    )

    assert response.status_code == 422
    issue = response.json()["detail"]["issues"][0]
    assert issue == {
        "worksheet": "单据导入",
        "row": 3,
        "column": "单据类型",
        "field": "document_type",
        "message": "Document type is not allowed for this import",
    }
    documents = client.get(
        f"{INVENTORY_PATH}/documents",
        headers=superuser_token_headers,
        params={"document_number": raw_number},
    )
    assert documents.status_code == 200
    assert documents.json()["count"] == 0


def test_excel_ledger_export_applies_document_page_filters(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    other_processing_unit = _create_processing_unit(client, superuser_token_headers)
    receiving_unit = _create_receiving_unit(client, superuser_token_headers)
    matched_number = f"XLSX-LEDGER-{uuid.uuid4()}"
    target_date = date(2026, 8, 3)
    _create_raw_receipt(
        client,
        superuser_token_headers,
        processing_unit_id=processing_unit["id"],
        item_name="匹配台账",
        item_code="LEDGER-MATCH",
        document_number=matched_number,
        business_date=target_date,
    )
    _create_raw_receipt(
        client,
        superuser_token_headers,
        processing_unit_id=processing_unit["id"],
        item_name="日期不匹配台账",
        item_code="LEDGER-DATE",
        document_number=f"XLSX-LEDGER-DATE-{uuid.uuid4()}",
        business_date=date(2026, 8, 4),
    )
    _create_raw_receipt(
        client,
        superuser_token_headers,
        processing_unit_id=other_processing_unit["id"],
        item_name="加工单位不匹配台账",
        item_code="LEDGER-UNIT",
        document_number=f"XLSX-LEDGER-UNIT-{uuid.uuid4()}",
        business_date=target_date,
    )

    filtered = client.get(
        f"{INVENTORY_PATH}/excel/ledger",
        headers=superuser_token_headers,
        params={
            "ledger_kind": "RAW",
            "processing_unit_id": processing_unit["id"],
            "document_number": matched_number,
            "business_date_from": str(target_date),
            "business_date_to": str(target_date),
        },
    )
    assert filtered.status_code == 200
    rows = list(
        load_workbook(BytesIO(filtered.content), read_only=True)["库存台账"].values
    )
    assert [row[2] for row in rows[1:]] == [matched_number]

    receiving_filtered = client.get(
        f"{INVENTORY_PATH}/excel/ledger",
        headers=superuser_token_headers,
        params={
            "ledger_kind": "RAW",
            "receiving_unit_id": receiving_unit["id"],
        },
    )
    assert receiving_filtered.status_code == 200
    receiving_rows = list(
        load_workbook(BytesIO(receiving_filtered.content), read_only=True)[
            "库存台账"
        ].values
    )
    assert len(receiving_rows) == 1


def test_excel_document_import_reports_inconsistent_document_groups(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    processing_unit = _create_processing_unit(client, superuser_token_headers)
    number = f"XLSX-CONFLICT-{uuid.uuid4()}"
    headers = [
        "单据类型",
        "日期",
        "单据号",
        "加工单位",
        "收货单位",
        "备注",
        "品名",
        "货号",
        "含毛量",
        "颜色",
        "缸号",
        "匹数",
        "米数",
    ]
    base_row = [
        "RAW_RECEIPT",
        "2026-08-01",
        number,
        processing_unit["name"],
        None,
        None,
        "冲突坯布",
        "CONFLICT-001",
        "100%",
        None,
        None,
        1,
        None,
    ]
    response = client.post(
        f"{INVENTORY_PATH}/excel/imports/documents",
        headers=superuser_token_headers,
        files={
            "workbook": (
                "conflict.xlsx",
                _xlsx_bytes(
                    [headers, base_row, [*base_row[:1], "2026-08-02", *base_row[2:]]]
                ),
                XLSX_MEDIA_TYPE,
            )
        },
    )

    assert response.status_code == 422
    assert response.json()["detail"]["issues"][0]["column"] == "日期"


def test_excel_legacy_import_and_ledger_export(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    raw_workbook = _xlsx_bytes(
        [
            ["日期", "加工单位", "品名", "品号", "含毛量", "入库"],
            [
                "2026-08-01",
                f"历史加工厂-{uuid.uuid4()}",
                "历史坯布",
                "H-001",
                "100%",
                2,
            ],
        ],
        sheet_name="历史坯布",
    )
    finished_workbook = _xlsx_bytes(
        [
            [
                "日期",
                "加工单位",
                "品名",
                "含毛量",
                "颜色+色号",
                "缸号",
                "入库匹数",
                "入库米数",
                "库存匹数",
                "库存米数",
            ],
            [
                "2026-08-01",
                f"历史成品加工厂-{uuid.uuid4()}",
                "历史成品",
                "70%",
                "焦糖",
                "LOT-001",
                2,
                20,
                2,
                20,
            ],
        ],
        sheet_name="历史成品",
    )
    imported = client.post(
        f"{INVENTORY_PATH}/excel/imports/legacy",
        headers=superuser_token_headers,
        files={
            "raw_workbook": ("raw.xlsx", raw_workbook, XLSX_MEDIA_TYPE),
            "finished_workbook": (
                "finished.xlsx",
                finished_workbook,
                XLSX_MEDIA_TYPE,
            ),
        },
    )

    assert imported.status_code == 201, imported.json()
    assert imported.json()["import_batch_id"]
    exported = client.get(
        f"{INVENTORY_PATH}/excel/ledger",
        headers=superuser_token_headers,
        params={"ledger_kind": "RAW"},
    )

    assert exported.status_code == 200
    worksheet = load_workbook(BytesIO(exported.content), read_only=True)["库存台账"]
    rows = list(worksheet.values)
    assert rows[0] == (
        "日期",
        "出入库类型",
        "单据号",
        "单位名称",
        "品名",
        "货号",
        "含毛量",
        "颜色",
        "缸号",
        "匹数变化",
        "米数变化",
        "备注",
    )
    assert any(row[4] == "历史坯布" and "历史来源" in row[11] for row in rows[1:])
    finished_export = client.get(
        f"{INVENTORY_PATH}/excel/ledger",
        headers=superuser_token_headers,
        params={"ledger_kind": "FINISHED"},
    )
    assert finished_export.status_code == 200
    finished_rows = list(
        load_workbook(BytesIO(finished_export.content), read_only=True)[
            "库存台账"
        ].values
    )
    assert any(row[4] == "历史成品" for row in finished_rows[1:])


def test_excel_legacy_import_reports_row_issues(
    client: TestClient, superuser_token_headers: dict[str, str]
) -> None:
    invalid_raw_workbook = _xlsx_bytes(
        [
            ["日期", "加工单位", "品名", "品号", "含毛量", "入库"],
            ["2026-08-01", "错误加工厂", "错误坯布", "E-001", "100%", "invalid"],
        ],
        sheet_name="错误坯布",
    )
    response = client.post(
        f"{INVENTORY_PATH}/excel/imports/legacy",
        headers=superuser_token_headers,
        files={
            "raw_workbook": ("raw.xlsx", invalid_raw_workbook, XLSX_MEDIA_TYPE),
            "finished_workbook": (
                "finished.xlsx",
                _xlsx_bytes([], sheet_name="Sheet"),
                XLSX_MEDIA_TYPE,
            ),
        },
    )

    assert response.status_code == 422
    issue = response.json()["detail"]["issues"][0]
    assert issue["worksheet"] == "错误坯布"
    assert issue["row"] == 2
    assert response.json()["request_id"]


def test_excel_endpoints_require_inventory_permissions(
    client: TestClient, normal_user_token_headers: dict[str, str]
) -> None:
    unauthenticated = client.get(f"{INVENTORY_PATH}/excel/templates/documents")
    forbidden = client.get(
        f"{INVENTORY_PATH}/excel/templates/documents",
        headers=normal_user_token_headers,
    )

    assert unauthenticated.status_code == 401
    assert forbidden.status_code == 403
