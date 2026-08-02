from io import BytesIO
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from app.core.excel import MAX_XLSX_DATA_ROWS, create_xlsx, read_xlsx_rows


class _Row(BaseModel):
    name: str = Field(alias="名称")
    quantity: int = Field(alias="数量")


def _workbook_bytes(rows: list[list[object]], *, sheet_name: str = "导入") -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_read_xlsx_rows_maps_aliases_and_ignores_unknown_columns() -> None:
    result = read_xlsx_rows(
        _workbook_bytes([["数量", "忽略", "名称"], [2, "x", "坯布"]]),
        model_type=_Row,
        worksheet_name="导入",
    )

    assert result.issues == []
    assert result.rows[0].row == 2
    assert result.rows[0].value == _Row(名称="坯布", 数量=2)


def test_read_xlsx_rows_reports_missing_and_duplicate_headers() -> None:
    result = read_xlsx_rows(
        _workbook_bytes([["名称", "名称"], ["坯布", "重复"]]),
        model_type=_Row,
        worksheet_name="导入",
    )

    assert [(issue.column, issue.message) for issue in result.issues] == [
        ("名称", "Duplicate declared header"),
        ("数量", "Missing declared header"),
    ]


def test_read_xlsx_rows_skips_blank_rows_and_locates_validation_errors() -> None:
    result = read_xlsx_rows(
        _workbook_bytes([["名称", "数量"], [None, None], ["坯布", "not-a-number"]]),
        model_type=_Row,
        worksheet_name="导入",
    )

    assert len(result.issues) == 1
    assert result.issues[0].row == 3
    assert result.issues[0].column == "数量"
    assert result.issues[0].field == "quantity"


def test_read_xlsx_rows_rejects_invalid_workbooks_and_limits() -> None:
    invalid = read_xlsx_rows(b"not a workbook", model_type=_Row, worksheet_name="导入")
    corrupted_output = BytesIO()
    with ZipFile(corrupted_output, "w") as archive:
        archive.writestr("not-a-workbook.txt", "corrupted")
    corrupted = read_xlsx_rows(
        corrupted_output.getvalue(), model_type=_Row, worksheet_name="导入"
    )
    missing_sheet = read_xlsx_rows(
        _workbook_bytes([["名称", "数量"]]), model_type=_Row, worksheet_name="缺失"
    )
    too_large = read_xlsx_rows(
        _workbook_bytes([["名称", "数量"]]),
        model_type=_Row,
        worksheet_name="导入",
        max_bytes=1,
    )
    too_many = read_xlsx_rows(
        _workbook_bytes(
            [["名称", "数量"]]
            + [[f"坯布-{index}", index] for index in range(MAX_XLSX_DATA_ROWS + 1)]
        ),
        model_type=_Row,
        worksheet_name="导入",
    )

    assert invalid.issues[0].message == "File is not a valid XLSX workbook"
    assert corrupted.issues[0].message == "File is not a valid XLSX workbook"
    assert "was not found" in missing_sheet.issues[0].message
    assert "byte limit" in too_large.issues[0].message
    assert "data row limit" in too_many.issues[0].message


def test_create_xlsx_uses_declared_alias_order() -> None:
    output = create_xlsx([_Row(名称="坯布", 数量=2)], model_type=_Row, worksheet_name="导入")
    worksheet = load_workbook(output, read_only=True, data_only=True)["导入"]

    assert list(worksheet.iter_rows(values_only=True)) == [("名称", "数量"), ("坯布", 2)]
