from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import func
from sqlmodel import Session, select

from app.core.exceptions import BadRequestError
from app.models import (
    InventoryDocumentLine,
    InventoryImportBatch,
    InventoryLedgerEntry,
    LegacyImportRow,
    ProcessingUnit,
    ReceivingUnit,
    User,
)
from app.models.inventory import InventoryMovementType
from app.modules.inventory.importer import import_workbooks


def _write_raw_workbook(path: Path, quantity: object) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["日期", "加工单位", "品名", "品号", "含毛量", "入库"])
    sheet.append(["2026-07-14", "回滚加工厂", "回滚坯布", "RB-001", "100%", quantity])
    workbook.save(path)


def test_import_rolls_back_batch_when_a_workbook_row_is_invalid(
    db: Session, tmp_path: Path
) -> None:
    actor = db.exec(select(User)).first()
    assert actor is not None
    raw_workbook = tmp_path / "raw-invalid.xlsx"
    finished_workbook = tmp_path / "finished-empty.xlsx"
    _write_raw_workbook(raw_workbook, "not-a-quantity")
    _write_raw_workbook(finished_workbook, 0)
    tracked_models = (
        InventoryDocumentLine,
        InventoryImportBatch,
        InventoryLedgerEntry,
        LegacyImportRow,
        ProcessingUnit,
        ReceivingUnit,
    )
    counts_before = {
        model: db.exec(select(func.count()).select_from(model)).one()
        for model in tracked_models
    }

    with pytest.raises(BadRequestError, match="Invalid inventory quantity"):
        import_workbooks(
            session=db,
            actor_user_id=actor.id,
            raw_workbook=raw_workbook,
            finished_workbook=finished_workbook,
        )

    assert {
        model: db.exec(select(func.count()).select_from(model)).one()
        for model in tracked_models
    } == counts_before
    assert (
        db.exec(
            select(ProcessingUnit).where(ProcessingUnit.normalized_name == "回滚加工厂")
        ).first()
        is None
    )


def test_import_rejects_an_inactive_human_without_creating_a_batch(
    db: Session, tmp_path: Path
) -> None:
    actor = User(
        email="inactive-importer@example.com",
        hashed_password="not-used",
        is_active=False,
    )
    db.add(actor)
    db.commit()
    batch_count_before = db.exec(
        select(func.count()).select_from(InventoryImportBatch)
    ).one()

    with pytest.raises(BadRequestError, match="must be active"):
        import_workbooks(
            session=db,
            actor_user_id=actor.id,
            raw_workbook=tmp_path / "not-read-raw.xlsx",
            finished_workbook=tmp_path / "not-read-finished.xlsx",
        )

    assert (
        db.exec(select(func.count()).select_from(InventoryImportBatch)).one()
        == batch_count_before
    )
    db.delete(actor)
    db.commit()


def test_import_accepts_a_preprovisioned_system_actor(
    db: Session, tmp_path: Path
) -> None:
    from app.core.audit import provision_system_actor

    actor = provision_system_actor(
        session=db,
        actor_key="inventory-test-import",
        email="inventory-test-import@system.invalid",
    )
    db.commit()
    raw_workbook = tmp_path / "raw-system-actor.xlsx"
    finished_workbook = tmp_path / "finished-system-actor.xlsx"
    _write_raw_workbook(raw_workbook, 0)
    Workbook().save(finished_workbook)

    import_workbooks(
        session=db,
        actor_user_id=actor.id,
        raw_workbook=raw_workbook,
        finished_workbook=finished_workbook,
    )

    batch = db.exec(
        select(InventoryImportBatch).order_by(InventoryImportBatch.imported_at.desc())
    ).first()
    assert batch is not None
    assert batch.created_by == actor.id
    assert batch.updated_by == actor.id


def test_import_reconciliation_opening_is_traceable_and_balances_by_key(
    db: Session, tmp_path: Path
) -> None:
    actor = db.exec(select(User)).first()
    assert actor is not None
    raw_workbook = tmp_path / "raw-return.xlsx"
    finished_workbook = tmp_path / "finished-empty.xlsx"
    _write_raw_workbook(raw_workbook, -2)
    _write_raw_workbook(finished_workbook, 0)

    report = import_workbooks(
        session=db,
        actor_user_id=actor.id,
        raw_workbook=raw_workbook,
        finished_workbook=finished_workbook,
    )

    batch = db.exec(
        select(InventoryImportBatch).order_by(InventoryImportBatch.imported_at.desc())
    ).first()
    assert batch is not None
    assert batch.reconciliation_report == report
    assert report["reconciliation_openings"] == 1

    entries = db.exec(
        select(InventoryLedgerEntry).where(
            InventoryLedgerEntry.import_batch_id == batch.id,
            InventoryLedgerEntry.item_name == "回滚坯布",
        )
    ).all()
    opening = next(
        entry
        for entry in entries
        if entry.movement_type is InventoryMovementType.MIGRATION_RECONCILIATION_OPENING
    )
    movement = next(
        entry
        for entry in entries
        if entry.movement_type is InventoryMovementType.RAW_RETURN
    )
    assert opening.document_line_id is None
    assert opening.legacy_import_row_id is not None
    assert opening.import_batch_id == batch.id
    assert opening.rolls_delta == Decimal("2")
    assert opening.reason == "历史迁移对账期初"
    assert movement.rolls_delta == Decimal("-2")
    assert sum(entry.rolls_delta for entry in entries) == Decimal("0")

    source = db.get(LegacyImportRow, opening.legacy_import_row_id)
    assert source is not None
    assert source.import_batch_id == batch.id
    assert source.requires_cleanup is False


def _write_finished_workbook(path: Path, rolls: float = 2) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "日期",
            "加工单位",
            "品名",
            "含毛量",
            "颜色+色号",
            "缸号",
            "入库匹数",
            "入库米数",
            "库存匹数",
            "库存米数",
        ]
    )
    sheet.append(
        [
            "2026-07-14",
            "颜色加工厂",
            "颜色成品",
            "70%",
            "焦糖",
            "LOT-001",
            rolls,
            30,
            2,
            30,
        ]
    )
    workbook.save(path)


def test_import_preserves_the_compound_color_column(
    db: Session, tmp_path: Path
) -> None:
    actor = db.exec(select(User)).first()
    assert actor is not None
    raw_workbook = tmp_path / "raw-empty.xlsx"
    finished_workbook = tmp_path / "finished-color.xlsx"
    _write_raw_workbook(raw_workbook, 0)
    _write_finished_workbook(finished_workbook)
    existing_import_batch_ids = set(db.exec(select(InventoryImportBatch.id)).all())

    import_workbooks(
        session=db,
        actor_user_id=actor.id,
        raw_workbook=raw_workbook,
        finished_workbook=finished_workbook,
    )

    new_import_batch_ids = set(db.exec(select(InventoryImportBatch.id)).all())
    new_import_batch_ids -= existing_import_batch_ids
    assert len(new_import_batch_ids) == 1
    import_batch_id = new_import_batch_ids.pop()
    ledger = db.exec(
        select(InventoryLedgerEntry).where(
            InventoryLedgerEntry.color_code == "焦糖",
            InventoryLedgerEntry.import_batch_id == import_batch_id,
        )
    ).one()
    assert ledger is not None
    assert ledger.color_code == "焦糖"


def test_import_preserves_decimal_rolls(db: Session, tmp_path: Path) -> None:
    actor = db.exec(select(User)).first()
    assert actor is not None
    raw_workbook = tmp_path / "raw-empty.xlsx"
    finished_workbook = tmp_path / "finished-decimal-rolls.xlsx"
    _write_raw_workbook(raw_workbook, 0)
    _write_finished_workbook(finished_workbook, rolls=0.5)

    import_workbooks(
        session=db,
        actor_user_id=actor.id,
        raw_workbook=raw_workbook,
        finished_workbook=finished_workbook,
    )

    line = db.exec(
        select(InventoryDocumentLine).where(
            InventoryDocumentLine.quantity_rolls == Decimal("0.5")
        )
    ).one()
    ledger = db.exec(
        select(InventoryLedgerEntry).where(
            InventoryLedgerEntry.document_line_id == line.id
        )
    ).one()
    assert line.quantity_rolls == Decimal("0.5")
    assert ledger.rolls_delta == Decimal("0.5")
