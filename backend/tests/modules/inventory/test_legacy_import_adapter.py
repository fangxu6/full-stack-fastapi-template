from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import CALENDAR_WINDOWS_1900

from app.core.exceptions import BadRequestError
from app.models.inventory import InventoryDocumentType
from app.modules.inventory.legacy_import_adapter import (
    _date,
    _movement_definitions,
    _row_values,
    _source_balance_snapshot,
    read_legacy_workbooks,
)

WORKSPACE_ROOT = Path(__file__).resolve().parents[4]


def _workbook_bytes(
    headers: list[object] | None = None, row: list[object] | None = None
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    if headers is not None:
        worksheet.append(headers)
    if row is not None:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@pytest.mark.parametrize(
    ("filename", "expected_row", "expected_values"),
    [
        (
            "2026年坯布入库明细.xlsx",
            3,
            {
                "品名": "驼坯",
                "品号": "2270-LTS-5",
                "库存": 5,
            },
        ),
        (
            "2026年成品进出明细.xlsx",
            4,
            {
                "品名": "染色双顺",
                "库存匹数": 12,
                "库存米数": 642.3,
            },
        ),
    ],
)
def test_row_values_recognizes_the_real_legacy_workbook_layouts(
    filename: str, expected_row: int, expected_values: dict[str, object]
) -> None:
    workbook_path = WORKSPACE_ROOT / "hongxia" / filename
    if not workbook_path.exists():
        pytest.skip(f"Legacy workbook fixture is unavailable: {workbook_path}")
    workbook = load_workbook(
        workbook_path,
        data_only=True,
        read_only=True,
    )

    source_row, cells = next(iter(_row_values(workbook["三泰"])))

    assert source_row == expected_row
    for field, expected_value in expected_values.items():
        assert cells[field] == expected_value


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (46090, date(2026, 3, 9)),
        ("46090", date(2026, 3, 9)),
        ("2026/3/9", date(2026, 3, 9)),
        ("2026-03-09 00:00:00", date(2026, 3, 9)),
        (datetime(2026, 3, 9), date(2026, 3, 9)),
        ("2025年结存", date(2025, 12, 31)),
    ],
)
def test_date_parser_preserves_legacy_workbook_dates(
    value: object, expected: date
) -> None:
    assert _date(value, epoch=CALENDAR_WINDOWS_1900) == expected


def test_date_parser_rejects_unrecognized_values() -> None:
    with pytest.raises(BadRequestError, match="Invalid inventory date"):
        _date(None, epoch=CALENDAR_WINDOWS_1900)


def test_raw_return_row_uses_its_base_processing_unit() -> None:
    movements, processing_unit_name = _movement_definitions(
        cells={"加工单位": "天双退走", "入库": 2},
        kind="RAW",
        worksheet_name="天双",
    )

    assert processing_unit_name == "天双"
    assert movements == [(InventoryDocumentType.RAW_RETURN, 2, None)]


def test_finished_row_keeps_roll_and_meter_snapshots() -> None:
    cells = {
        "入库匹数": 3,
        "入库米数": 120.5,
        "出库匹数": 1,
        "出库米数": 38.2,
        "库存匹数": 12,
        "库存米数": 642.3,
    }

    movements, _ = _movement_definitions(
        cells=cells,
        kind="FINISHED",
        worksheet_name="三泰",
    )

    assert movements == [
        (
            InventoryDocumentType.FINISHED_RECEIPT,
            Decimal("3"),
            Decimal("120.5"),
        ),
        (
            InventoryDocumentType.FINISHED_SHIPMENT,
            Decimal("1"),
            Decimal("38.2"),
        ),
    ]
    assert _source_balance_snapshot(cells, "FINISHED") == {
        "rolls": 12,
        "meters": 642.3,
    }


def test_finished_row_preserves_decimal_rolls() -> None:
    movements, _ = _movement_definitions(
        cells={"出库匹数": 0.5, "出库米数": 20},
        kind="FINISHED",
        worksheet_name="三泰",
    )

    assert movements == [
        (
            InventoryDocumentType.FINISHED_SHIPMENT,
            Decimal("0.5"),
            Decimal("20"),
        )
    ]


def test_negative_finished_quantity_is_normalized_as_a_shipment() -> None:
    movements, _ = _movement_definitions(
        cells={"入库匹数": -1, "入库米数": -35.5},
        kind="FINISHED",
        worksheet_name="三泰",
    )

    assert movements == [
        (
            InventoryDocumentType.FINISHED_SHIPMENT,
            Decimal("1"),
            Decimal("35.5"),
        )
    ]


def test_meter_only_finished_history_is_preserved() -> None:
    movements, _ = _movement_definitions(
        cells={"入库匹数": 0, "入库米数": 20},
        kind="FINISHED",
        worksheet_name="三泰",
    )

    assert movements == [
        (
            InventoryDocumentType.FINISHED_RECEIPT,
            Decimal("0"),
            Decimal("20"),
        )
    ]


def test_legacy_adapter_returns_normalized_records() -> None:
    raw_content = _workbook_bytes(
        ["日期", "加工单位", "品名", "品号", "含毛量", "入库", "库存"],
        ["2026-08-07", "加工厂", "坯布", "RAW-1", "100%", 2, 5],
    )
    empty_content = _workbook_bytes()

    records, issues = read_legacy_workbooks(
        raw_content=raw_content,
        raw_filename="raw.xlsx",
        finished_content=empty_content,
        finished_filename="finished.xlsx",
    )

    assert issues == []
    assert len(records) == 1
    assert records[0].item_name == "坯布"
    assert len(records[0].movements) == 1
    assert records[0].movements[0].document_type is InventoryDocumentType.RAW_RECEIPT
    assert records[0].movements[0].rolls == Decimal("2")
    assert records[0].source_balance_snapshot == {"rolls": 5}


def test_legacy_adapter_reports_malformed_workbook_rows() -> None:
    raw_content = _workbook_bytes(
        ["日期", "加工单位", "品名", "品号", "含毛量", "入库"],
        ["not-a-date", "加工厂", "坯布", "RAW-1", "100%", 2],
    )
    empty_content = _workbook_bytes()

    records, issues = read_legacy_workbooks(
        raw_content=raw_content,
        raw_filename="raw.xlsx",
        finished_content=empty_content,
        finished_filename="finished.xlsx",
    )

    assert records == []
    assert len(issues) == 1
    assert issues[0].row == 2
    assert issues[0].message == "Invalid inventory date: not-a-date"
